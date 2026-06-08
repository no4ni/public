#!/usr/bin/env python3
"""
sensor.py
Превращает любой файл в текстовое описание.
Использует различные библиотеки для анализа изображений, аудио, видео, документов.
"""

import os
import sys
import json
import hashlib
import math
import warnings
from pathlib import Path

# Сторонние библиотеки (необходимо установить)
try:
	import magic  # pip install python-magic-bin
except ImportError:
	magic = None

try:
	from PIL import Image, ImageEnhance, ImageFilter, ExifTags
except ImportError:
	Image = None

try:
	import pytesseract
except ImportError:
	pytesseract = None

try:
	import openpyxl
except ImportError:
	openpyxl = None

try:
	import librosa
	import soundfile as sf
except ImportError:
	librosa = None

try:
	import cv2
except ImportError:
	cv2 = None

try:
	from pdf2image import convert_from_path
except ImportError:
	convert_from_path = None

try:
	import docx
except ImportError:
	docx = None

# ---------- Лимит ----------
MAX_TEXT_SIZE = 6144  # байт

# ---------- Цвета ----------
COLORS = []

def load_colors(filename=None):
	"""Загружает цвета из файла с табуляцией: название, hex, r, g, b."""
	if filename is None:
		filename = Path(__file__).parent / "colors.txt"
	colors = []
	try:
		with open(filename, 'r', encoding='utf-8') as f:
			for line in f:
				line = line.strip()
				if not line or line.startswith('#'):
					continue
				parts = line.split('\t')
				if len(parts) >= 5:
					name = parts[0]
					try:
						r = int(parts[2])
						g = int(parts[3])
						b = int(parts[4])
						colors.append((name, r, g, b))
					except ValueError:
						continue
	except FileNotFoundError:
		print("Файл с цветами не найден, буду выводить только RGB", file=sys.stderr)
	return colors

def closest_color(r, g, b, colors):
	"""Возвращает название ближайшего цвета по евклидову расстоянию."""
	if not colors:
		return None
	min_dist = float('inf')
	best_name = None
	for name, cr, cg, cb in colors:
		dist = (r - cr)**2 + (g - cg)**2 + (b - cb)**2
		if dist < min_dist:
			min_dist = dist
			best_name = name
	return best_name

# ---------- Обработчики файлов ----------
def handle_text(filepath):
	"""Читает текстовый файл, обрезает до лимита."""
	try:
		with open(filepath, 'r', encoding='utf-8') as f:
			content = f.read()
	except UnicodeDecodeError:
		try:
			with open(filepath, 'r', encoding='cp1251') as f:
				content = f.read()
		except Exception as e:
			content = f"Не удалось прочитать файл как текст: {e}"
	except Exception as e:
		return f"Ошибка чтения текста: {e}"

	if len(content) > MAX_TEXT_SIZE:
		content = content[:MAX_TEXT_SIZE] + "\n... (обрезано)"
	return f"--- Текстовый файл ---\n{content}"

def handle_image(filepath, config, lang='rus+eng'):
	"""Извлекает метаданные, цвета, OCR-текст."""
	if Image is None or pytesseract is None:
		return "Для обработки изображений нужны Pillow и pytesseract."

	try:
		original = Image.open(filepath)
	except Exception as e:
		return f"Ошибка открытия изображения: {e}"

	# Подготовка изображения для OCR
	img = original.copy()
	if img.mode != 'L':
		img = img.convert('L')
	enhancer = ImageEnhance.Contrast(img)
	img = enhancer.enhance(2.0)

	width, height = img.size
	if width < 1500:
		scale = 2000 / width
		new_size = (int(width * scale), int(height * scale))
		img = img.resize(new_size, Image.Resampling.LANCZOS)

	try:
		ocr_text = pytesseract.image_to_string(img, lang=lang, config=config).strip()
	except Exception as e:
		ocr_text = f"OCR ошибка: {e}"

	if len(ocr_text) > MAX_TEXT_SIZE:
		ocr_text = ocr_text[:MAX_TEXT_SIZE] + "\n... (обрезано)"

	# EXIF
	exif = {}
	if hasattr(original, '_getexif') and original._getexif():
		exif_data = original._getexif()
		for tag, value in exif_data.items():
			decoded = ExifTags.TAGS.get(tag, tag)
			exif[decoded] = str(value)[:100]

	# Средний цвет
	if original.mode == 'RGBA':
		rgb_original = original.convert('RGB')
	else:
		rgb_original = original

	with warnings.catch_warnings():
		warnings.simplefilter("ignore", DeprecationWarning)
		pixels = list(rgb_original.getdata())

	total_pixels = len(pixels)
	info = {"size": original.size}
	if total_pixels > 0:
		r_total = sum(p[0] for p in pixels)
		g_total = sum(p[1] for p in pixels)
		b_total = sum(p[2] for p in pixels)
		avg_r = int(r_total / total_pixels)
		avg_g = int(g_total / total_pixels)
		avg_b = int(b_total / total_pixels)
		if COLORS:
			nearest = closest_color(avg_r, avg_g, avg_b, COLORS)
			if nearest:
				info["closest_color"] = nearest
		else:
			info["avg_color"] = f"R:{avg_r} G:{avg_g} B:{avg_b}"

	result = f"--- Изображение {filepath} ---\n"
	result += json.dumps(info, indent=2, ensure_ascii=False) + "\n"
	if exif:
		result += "EXIF:\n" + json.dumps(exif, indent=2, ensure_ascii=False) + "\n"
	if ocr_text:
		result += "OCR текст:\n" + ocr_text + "\n"

	return result[:MAX_TEXT_SIZE]

def handle_audio(filepath):
	"""Извлекаем основные характеристики аудио."""
	if librosa is None:
		return "Для аудио нужны librosa и soundfile: pip install librosa soundfile"
	try:
		y, sr = librosa.load(filepath, sr=None, mono=True)
		duration = librosa.get_duration(y=y, sr=sr)
		rms = librosa.feature.rms(y=y).mean()
		spec_cent = librosa.feature.spectral_centroid(y=y, sr=sr).mean()
		tempo, _ = librosa.beat.beat_track(y=y, sr=sr)
		result = f"--- Аудиофайл ---\n"
		result += f"Длительность: {duration:.2f} сек\n"
		result += f"Частота дискретизации: {sr} Гц\n"
		result += f"Средняя громкость (RMS): {rms:.4f}\n"
		result += f"Спектральный центроид: {spec_cent:.2f} Гц\n"
		result += f"Предполагаемый темп: {tempo:.1f} BPM\n"
		return result[:MAX_TEXT_SIZE]
	except Exception as e:
		return f"Ошибка обработки аудио: {e}"

def handle_video(filepath):
	"""Извлекаем информацию о видео."""
	if cv2 is None:
		return "Для видео нужен opencv-python: pip install opencv-python"
	cap = cv2.VideoCapture(filepath)
	if not cap.isOpened():
		return "Не удалось открыть видео"
	info = {
		"frame_width": int(cap.get(cv2.CAP_PROP_FRAME_WIDTH)),
		"frame_height": int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT)),
		"fps": cap.get(cv2.CAP_PROP_FPS),
		"total_frames": int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
	}
	info["duration"] = info["total_frames"] / info["fps"] if info["fps"] else 0
	ret, frame = cap.read()
	cap.release()
	result = f"--- Видеофайл ---\n"
	if ret:
		result += "Первый кадр успешно извлечён (анализ не проводится).\n"
	result += json.dumps(info, indent=2, ensure_ascii=False)
	return result[:MAX_TEXT_SIZE]

def handle_pdf(filepath, lang='rus+eng'):
	"""OCR PDF через Tesseract."""
	if convert_from_path is None or pytesseract is None:
		return "Для OCR PDF нужны pdf2image и pytesseract."
	try:
		images = convert_from_path(filepath, dpi=300)
	except Exception as e:
		return f"Ошибка конвертации PDF: {e}"

	full_text = ""
	total_pages = len(images)
	print(f"Начинаю OCR для PDF, страниц: {total_pages}", file=sys.stderr)
	for i, image in enumerate(images):
		print(f"Обработка страницы {i+1}...", file=sys.stderr)
		try:
			text = pytesseract.image_to_string(image, lang=lang, config='--psm 3')
			full_text += text + "\n\n"
		except Exception as e:
			full_text += f"[Ошибка OCR на странице {i+1}: {e}]\n\n"
		if len(full_text) > MAX_TEXT_SIZE:
			full_text = full_text[:MAX_TEXT_SIZE] + "\n... (обрезано, превышен лимит)"
			break
	return f"--- PDF документ (OCR Tesseract) ---\n{full_text}"

def handle_docx(filepath):
	"""Извлечение текста из DOCX."""
	if docx is None:
		return "Для DOCX нужен python-docx: pip install python-docx"
	try:
		doc = docx.Document(filepath)
		text = [para.text for para in doc.paragraphs]
		full_text = "\n".join(text)
		if len(full_text) > MAX_TEXT_SIZE:
			full_text = full_text[:MAX_TEXT_SIZE] + "\n... (обрезано)"
		return f"--- DOCX документ ---\n{full_text}"
	except Exception as e:
		return f"Ошибка чтения DOCX: {e}"

def handle_excel(filepath):
	"""Извлекает текст из Excel."""
	if openpyxl is None:
		return "Для Excel нужна библиотека openpyxl: pip install openpyxl"
	try:
		wb = openpyxl.load_workbook(filepath, data_only=True)
		all_text = []
		for sheetname in wb.sheetnames:
			ws = wb[sheetname]
			sheet_text = f"--- Лист: {sheetname} ---\n"
			for row in ws.iter_rows(values_only=True):
				if all(cell is None for cell in row):
					continue
				row_str = "\t".join(str(cell) if cell is not None else "" for cell in row)
				sheet_text += row_str + "\n"
			all_text.append(sheet_text)
			total_len = sum(len(t) for t in all_text)
			if total_len > MAX_TEXT_SIZE:
				break
		full_text = "\n".join(all_text)
		if len(full_text) > MAX_TEXT_SIZE:
			full_text = full_text[:MAX_TEXT_SIZE] + "\n... (обрезано)"
		return f"--- Excel документ ---\n{full_text}"
	except Exception as e:
		return f"Ошибка чтения Excel: {e}"

def handle_binary(filepath):
	"""Для неизвестных типов: базовая информация, энтропия, строки."""
	size = os.path.getsize(filepath)
	mime = magic.from_file(filepath, mime=True) if magic else "unknown"
	desc = magic.from_file(filepath) if magic else "unknown"
	sha256 = hashlib.sha256()
	with open(filepath, 'rb') as f:
		for chunk in iter(lambda: f.read(65536), b''):
			sha256.update(chunk)
	file_hash = sha256.hexdigest()

	# Энтропия первых 4K
	entropy = 0
	if size > 0:
		counts = {}
		with open(filepath, 'rb') as f:
			data = f.read(4096)
		for byte in data:
			counts[byte] = counts.get(byte, 0) + 1
		entropy = -sum((c/len(data)) * math.log2(c/len(data)) for c in counts.values())

	# Поиск ASCII-строк
	strings = []
	with open(filepath, 'rb') as f:
		data = f.read(4096)
		current = []
		for b in data:
			if 32 <= b <= 126:
				current.append(chr(b))
			else:
				if len(current) >= 4:
					strings.append(''.join(current))
				current = []
		if len(current) >= 4:
			strings.append(''.join(current))

	result = f"--- Неизвестный/бинарный файл ---\n"
	result += f"Размер: {size} байт\n"
	result += f"MIME-тип: {mime}\n"
	result += f"Описание: {desc}\n"
	result += f"SHA256: {file_hash}\n"
	result += f"Энтропия первых 4K: {entropy:.4f}\n"
	if strings:
		result += "Найденные строки (первые 10):\n" + "\n".join(strings[:10]) + "\n"
	return result[:MAX_TEXT_SIZE]

# ---------- Диспетчер ----------
def process_file(filepath, config, lang='rus+eng'):
	path = Path(filepath)
	if not path.exists():
		return "Файл не найден."
	ext = path.suffix.lower()

	# По расширению
	if ext in ['.txt', '.log', '.py', '.js', '.html', '.css', '.json', '.xml', '.csv']:
		return handle_text(filepath)
	elif ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.webp']:
		return handle_image(filepath, config, lang)
	elif ext in ['.xlsx', '.xls', '.xlsm']:
		return handle_excel(filepath)
	elif ext in ['.mp3', '.wav', '.flac', '.aac', '.ogg', '.m4a']:
		return handle_audio(filepath)
	elif ext in ['.mp4', '.avi', '.mov', '.mkv', '.webm']:
		return handle_video(filepath)
	elif ext == '.pdf':
		return handle_pdf(filepath, lang)
	elif ext in ['.docx', '.doc']:
		return handle_docx(filepath)
	else:
		# По MIME
		if magic:
			try:
				mime = magic.from_file(filepath, mime=True)
				if mime.startswith('text/'):
					return handle_text(filepath)
				elif mime.startswith('image/'):
					return handle_image(filepath, config, lang)
				elif mime.startswith('audio/'):
					return handle_audio(filepath)
				elif mime.startswith('video/'):
					return handle_video(filepath)
				elif mime == 'application/pdf':
					return handle_pdf(filepath, lang)
				elif mime in ['application/msword', 'application/vnd.openxmlformats-officedocument.wordprocessingml.document']:
					return handle_docx(filepath)
				elif mime in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.ms-excel']:
					return handle_excel(filepath)
			except:
				pass
		return handle_binary(filepath)

# ---------- Точка входа ----------
if __name__ == '__main__':
	MAX_TEXT_SIZE = 6144
	COLORS = load_colors()

	output_file = None
	config = None
	filepath = None
	lang = 'rus+eng'

	# Разбор аргументов
	args = sys.argv[1:]
	i = 0
	while i < len(args):
		arg = args[i]
		if arg == '--nolimit':
			MAX_TEXT_SIZE = 30 * 1024 * 1024
		elif arg == '--output':
			if i + 1 < len(args):
				output_file = args[i + 1]
				i += 1
			else:
				print("Ошибка: после --output необходимо указать имя файла.", file=sys.stderr)
				sys.exit(1)
		elif arg == '--config':
			if i + 1 < len(args):
				config = args[i + 1]
				i += 1
			else:
				print("Ошибка: после --config необходимо указать строку конфигурации.", file=sys.stderr)
				sys.exit(1)
		elif arg.startswith('-'):
			print(f"Неизвестный флаг: {arg}", file=sys.stderr)
			sys.exit(1)
		else:
			if filepath is None:
				filepath = arg
		i += 1

	if filepath is None:
		print("Использование: python sensor.py <путь_к_файлу> [язык] [--nolimit] [--output файл] [--config строка]")
		sys.exit(1)

	output = process_file(filepath, config, lang)

	if output_file:
		try:
			with open(output_file, 'w', encoding='utf-8') as f:
				f.write(output)
			print(f"Результат сохранён в файл: {output_file}", file=sys.stderr)
		except Exception as e:
			print(f"Ошибка при записи в файл: {e}", file=sys.stderr)
	else:
		try:
			print(output)
		except UnicodeEncodeError:
			print("\nОшибка кодировки терминала. Используйте --output <файл> для сохранения результата.", file=sys.stderr)